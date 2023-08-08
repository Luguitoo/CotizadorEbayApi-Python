#imports
import PySimpleGUI as sg
#Web scraping para cotizacion del dolar del dia
import requests
from bs4 import BeautifulSoup
#EbaySDK
from ebaysdk.finding import Connection
import datetime
from deep_translator import GoogleTranslator
#Pillow para crea imagenes
from PIL import Image #Para cargar las imagenes
from PIL import ImageDraw #para dibujar
from PIL import ImageOps #Reescalado
from PIL import ImageFont #Fuentes
#Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

from datetime import datetime

r = requests.get('https://www.cambioschaco.com.py') #Url de donde obtener los datos
soup = BeautifulSoup(r.text, 'html.parser')
cotizacion = soup.find_all('span', class_="sale", limit = 1) #Dolar
#print(cotizacion[0].text.replace('.', ""))
dolar = float(cotizacion[0].text.replace('.', ""))

API_KEY = ""

#Leer la key del txt.. "key.txt"
with open('key.txt', 'r') as f:
    API_KEY = f.readline().strip()
    #print(API_KEY)

#Leer la config del user
def read_config(file_path):
    config_data = {}

    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith('Profits:'):
                config_data['Profits'] = int(line.split(':')[1].strip())
                print(int(line.split(':')[1].strip()))
            elif line.startswith('Shipping:'):
                config_data['Shipping'] = int(line.split(':')[1].strip())
            elif line.startswith('Template:'):
                config_data['Template'] = int(line.split(':')[1].strip())

    return config_data

config_data = read_config("config.txt")
print("Configuración leída:")
print("Profits:", config_data.get('Profits'))
print("Shipping:", config_data.get('Shipping'))
print("Template:", config_data.get('Template'))

#Guardar los datos del config.txt
def update_config_file(ganancias, envio, tema):
    # Leer el contenido del archivo
    with open("config.txt", "r") as file:
        content = file.readlines()

    # Actualizar los valores en el contenido
    updated_content = []
    for line in content:
        if "Profits:" in line:
            line = f"Profits: {ganancias}\n"
        elif "Shipping:" in line:
            line = f"Shipping: {envio}\n"
        elif "Template:" in line:
            line = f"Template: {tema}\n"
        updated_content.append(line)

    # Escribir el contenido actualizado nuevamente en el archivo
    with open("config.txt", "w") as file:
        file.writelines(updated_content)

class Ebay(object):
    def __init__(self, API_KEY, keyword, capacity, resultado):
        self.api_key = API_KEY
        self.keyword = keyword
        self.capacity = capacity
        self.resultado = resultado
    def fetch(self):
        try:
            api = Connection(appid=self.api_key, config_file=None)
            #conditionId 2020 es Very Good
            #response = api.execute('findItemsAdvanced', {'keywords': 'Iphone X 64GB Very Good', 'country':'US', 'condition.conditionId':'2020', 'capacity':'64'})
            response = api.execute('findItemsAdvanced', {'keywords': self.keyword +' '+self.capacity + 'GB', 'country':'US', 'condition': {'conditionId': '2020'}})
            #print(response) #Comprobar que esta conectando recibindo un objeto
            #print(response.reply) #Resultados con la busqueda del keyword mandado, tambien sirve para ver los objetos disponibles
            #item = response.reply.searchResult.item[0]
            cont = 0
            lista = []
            for item in response.reply.searchResult.item: #De los resultados, imprime el titulo y el precio actual.
                if item.condition.conditionId == "2020":
                    #print(item)
                    #Todos los resultados
                    x = self.keyword.split("Pro")
                    x1 = self.keyword.split("Max")
                    if len(x1) > 1 or len(x) > 1:
                        if len(x1) > 1 and len(x) > 1:
                            c = item.title.split("Pro Max")
                            if len(c) > 1:
                                #print(f"Title: {item.title}, Price {item.sellingStatus.currentPrice.value}, Condition {item.condition.conditionId}, Capacidad {item.capacity}")
                                lista.append(item)
                        if len(x) > 1:
                            c = item.title.split("Pro Max")
                            if len(c) == 1:
                                #print(f"Title: {item.title}, Price {item.sellingStatus.currentPrice.value}, Condition {item.condition.conditionId}")
                                lista.append(item)
                    else:
                        x = item.title.split("Pro Max")
                        x1 = item.title.split("Pro")
                        if len(x1) == 1 and len(x) == 1:
                            #print(f"Title: {item.title}, Price {item.sellingStatus.currentPrice.value}, Condition {item.condition.conditionId}")
                            lista.append(item)
            resultado = []
            for item in lista:
                x = item.title.split("GB")
                x2 = item.title.split(f"{self.capacity}")
                if len(x) == 2 and len(x2) > 1:
                    resultado.append(item)
            #print("Lista filtrada: ")
            for item in resultado:
                x = item.title.split("mini")
                if len(x) == 1:
                    #print(f"Title: {item.title}, Price {item.sellingStatus.currentPrice.value}, Condition {item.condition.conditionId}")
                    if cont == 0:
                        self.resultado = item
                        cont += 1

        except ConnectionError as e:
            print(e)
            print(e.response.dict())
    def parse(self):
        pass  
    
#Convertir el precio a un numero con puntos y redondeado
def pre_puntos(precio):
    a = list(precio)
    nr = ""
    cont = len(precio)
    for j in range(0, len(precio)):
        if j > 3:
            nr = nr + "0"
        else:
            nr = nr + a[j]
        cont -= 1
        if cont % 3 == 0 and j != len(precio) - 1:
            nr = nr + "."
    #print(nr)
    return nr  
#Convertir los productos a imagenes
#Convertir el precio a un numero con puntos y redondeado
def pre_puntos(precio):
    a = list(precio)
    nr = ""
    cont = len(precio)
    for j in range(0, len(precio)):
        if j > 3:
            nr = nr + "0"
        else:
            nr = nr + a[j]
        cont -= 1
        if cont % 3 == 0 and j != len(precio) - 1:
            nr = nr + "."
    #print(nr)
    return nr


def conv_img(nombre, precios):
    #Crea una imagen del producto con su precio...
    print(f"Procesando: {nombre}")
    dir_img = "./imagenes/"
    dir_rec = "./recursos/"
    dir_save = "./post/"
    #Background
    if int(config_data.get('Template')) == 1:
        image = Image.open(dir_img + f"{nombre}-Orange.png")
    elif int(config_data.get('Template')) == 2:
        image = Image.open(dir_img + f"{nombre}-Yellow.png")
    elif int(config_data.get('Template')) == 3:
        image = Image.open(dir_img + f"{nombre}-Green.png")
    draw = ImageDraw.Draw(image)
    print(precios)

    #Escribimos precio y titulo #500, 900
    precio1 = precios[0] + " Gs"
    precio2 = precios[1] + " Gs"
    if len(precios) == 2:
        fuente = ImageFont.truetype(dir_rec + "Antonio-Bold.ttf", size=67, encoding="UTF-8")
        ancho, alto = draw.textsize(precio1, font=fuente)
        draw.text((800 - ancho // 2, 318 - alto // 2), precio1, font=fuente,fill=(39, 39, 39))
        draw.text((805 - ancho // 2, 684 - alto // 2), precio2, font=fuente,fill=(39, 39, 39))
    else:
        precio3 = precios[2] + " Gs"
        fuente = ImageFont.truetype(dir_rec + "Antonio-Bold.ttf", size=60, encoding="UTF-8")
        ancho, alto = draw.textsize(precio1, font=fuente)
        draw.text((800 - ancho // 2, 267 - alto // 2), precio1, font=fuente,fill=(39, 39, 39))
        draw.text((805 - ancho // 2, 538 - alto // 2), precio2, font=fuente,fill=(39, 39, 39))
        draw.text((805 - ancho // 2, 775 - alto // 2), precio3, font=fuente,fill=(39, 39, 39))
    #guardamos la imagen del producto
    image.save(dir_save + nombre + ".png")
    print("Success")
    precios = []
    return 0
#tema simplegui
##sg.theme("reddit")
# Definir el esquema de colores naranjas
my_orange_theme = {
    'BACKGROUND': 'white',
    'TEXT': 'black',
    'INPUT': '#FFFFFF',
    'TEXT_INPUT': 'black',
    'SCROLL': '#FFD700',
    'BUTTON': ('white', '#FFA500'),
    'PROGRESS': ('#01826B', '#D0D0D0'),
    'BORDER': 1,
    'SLIDER_DEPTH': 0,
    'PROGRESS_DEPTH': 0,
}

# Agregar el esquema de colores a la lista de temas
sg.theme_add_new('MyOrangeTheme', my_orange_theme)

# Establecer el tema predeterminado
sg.theme('MyOrangeTheme')
#Cabecera, Productos
headings= ['Producto', 'Capacidad','Costo', 'Venta', 'Condición']
data = []
def make_win1():
    # Columna 1:
    col1 = [[sg.Image(filename='./imagenes/logo.png' , size=(100, 80))],
            [sg.Text("Cotización del dia")],
            [sg.Text(f"Dolar: {cotizacion[0].text} Gs")],
            [sg.Text("Producto:"), sg.Combo(["Iphone X", "Iphone 11", "Iphone 12", "Iphone 13"], default_value="Iphone X" , key='-COMBO-')],
            [sg.Button('Buscar'), sg.Button('Borrar'),],
            ]
     # Columna 2:

    col2 = [[sg.Button('Opciones', key='-SETTINGS-')],
            [sg.Table(data, headings=headings, justification='left', key='ta_p', enable_events=True, col_widths=[60,60, 60, 60, 60])],
            [sg.Text("Costo total:  "), sg.Text("0",key='ttc'), sg.Text("Ganancias:  "), sg.Text("0",key='ttg'), sg.Button('Exportar presupusto'),]]

    layout = [[sg.Column(col1), sg.Column(col2, element_justification='right')]]

    return sg.Window('Cotizaciones de productos', layout, location=(800,600), finalize=True)

def make_win2():
    #temas
    if int(config_data.get('Template')) == 1:
        tema_imagenes = [[sg.Radio("Naranja ", "tema", key="Naranja", default=True), sg.Image(filename='imagenes/preview1.png', size=(100, 80))],
                        [sg.Radio("Amarillo", "tema", key="Amarillo",), sg.Image(filename='imagenes/preview2.png', size=(100, 80))],
                        [sg.Radio("Verde   ", "tema", key="Verde",), sg.Image(filename='imagenes/preview3.png', size=(100, 80))],
                        ]
    elif int(config_data.get('Template')) == 2:
        tema_imagenes = [[sg.Radio("Naranja ", "tema", key="Naranja", ), sg.Image(filename='imagenes/preview1.png', size=(100, 80))],
                        [sg.Radio("Amarillo", "tema", key="Amarillo", default=True), sg.Image(filename='imagenes/preview2.png', size=(100, 80))],
                        [sg.Radio("Verde   ", "tema", key="Verde",), sg.Image(filename='imagenes/preview3.png', size=(100, 80))],
                        ]
    elif int(config_data.get('Template')) == 3:
        tema_imagenes = [[sg.Radio("Naranja ", "tema", key="Naranja", ), sg.Image(filename='imagenes/preview1.png', size=(100, 80))],
                        [sg.Radio("Amarillo", "tema", key="Amarillo", ), sg.Image(filename='imagenes/preview2.png', size=(100, 80))],
                        [sg.Radio("Verde   ", "tema", key="Verde", default=True), sg.Image(filename='imagenes/preview3.png', size=(100, 80))],
                        ]
    layout = [
            [sg.Text("Opciones:")],
            [sg.Text("Ganacias:"), sg.Input(key='ganancias', size=(20, 1), default_text=config_data.get('Profits')), sg.Text("USD")],
            [sg.Text("Costo envio:"), sg.Input(key='ship', size=(20, 1), default_text=config_data.get('Shipping')), sg.Text("USD")],
            [sg.Text("Tema de imagenes: ")],
            *tema_imagenes,  # Utilizamos el operador * para desempaquetar la lista de opciones del tema
            [sg.Button('Guardar'),sg.Button('Cerrar')]
        ]

    return sg.Window('Segunda Ventana', layout, finalize=True)


# Definimos una lista para almacenar los productos
productos = []
precios = []
#Actualizacion tabla de productos
def act():
    window["ta_p"].update(values=productos)
    window.refresh()

def open_win2():
    window2 = make_win2()
    while True:
        event, values = window2.read()
        if event == sg.WIN_CLOSED or event == 'Cerrar':
            break
        elif event == "Guardar":
            gana = int((values['ganancias']))
            env = int((values['ship']))
            #Obtener el valor seleccionado en el grupo de radio buttons
            if values["Naranja"]:
                tema = 1
            elif values["Amarillo"]:
                tema = 2
            elif values["Verde"]:
                tema = 3
            print(env, gana, tema) 
            # Actualizar el archivo de configuración con los nuevos valores
            update_config_file(gana, env, tema)
            #Actualiza los datos de la sesion
            config_data['Profits'] = gana
            config_data['Shipping'] = env
            config_data['Template'] = tema
            break
    window2.close()
window1, window2 = make_win1(), None

costo_t = 0
gasto_t = 0

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WIN_CLOSED or event == 'Salir':
        window.close()
        ##if window == window2:       # if closing win 2, mark as closed
            ##window2 = None
        if window == window1: # if closing win 1, exit program
            break 
    elif event == '-SETTINGS-':
        open_win2()   
    elif event == 'Buscar' :
        pro = values['-COMBO-'].lower()   #minisculas
        capacidades = []
        if pro == "iphone x":
                capacidades = ["64", "256"]
        if pro == "iphone 11":
                capacidades = ["64","128","256"]
        if pro == "iphone 12":
                capacidades = ["64","128", "256"]
        if pro == "iphone 13":
                capacidades = ["128","256"]
        ##print(capacidades)
        for x in range(0, len(capacidades)):
                #Conexion a la API y busqueda del celular
                e = Ebay(API_KEY, pro, capacidades[x], "")
                e.fetch()
                e.parse()
                #Mejor resultado
                busqueda = e.resultado
                #print(busqueda)
                # Tiene 7% de taxes  y 50mil cuesta el envio, recargar eso
                precio = str(int(float(busqueda.sellingStatus.currentPrice.value) * dolar))
                preciov = str((int((float(busqueda.sellingStatus.currentPrice.value) + config_data.get('Profits') + config_data.get('Shipping')) * dolar)))
                # print(precio)
                a = list(precio)
                # print(precio)
                #print(preciov)
                preciob = pre_puntos(precio)
                preciov = pre_puntos(preciov)
                print("Mejor elección: ")
                titulo = str(busqueda.title)
                # print(titulo)
                traductor = GoogleTranslator(source='en', target='es')
                titulo = traductor.translate(busqueda.title)
                condicion = traductor.translate(busqueda.condition.conditionDisplayName).split(" - ")
                condicion = condicion[0]
                # print(titulo)
                print(f"Titulo: {busqueda.title}, Precio de Ebay: {preciob}, Precio de venta: {preciov} Gs, Condicion: {condicion}")
                # Agregamos eñ producto a la lista
                nombre = busqueda.title.split()
                producto = [nombre[1] + " " + nombre[2],capacidades[x] + " GB",preciob, preciov, condicion]
                productos.append(producto)
                precios.append(preciov)
                #Actualizar valores de costos y ganancias
                preciob = preciob.replace('.', '')       # Eliminar los puntos de la cadena
                costo_t += float(preciob)
                costo = '{:,.0f}'.format(costo_t)
                window1['ttc'].update(costo)
                preciov = preciov.replace('.', '')
                gasto_t += float(preciov)
                gasto = '{:,.0f}'.format(gasto_t)
                window1['ttg'].update(gasto)
                print(productos)
                act()
                producto = []
        #Convierte los precios obtenidos en imagen para postear
        conv_img(pro, precios)
    elif event == 'Exportar presupusto':
        print("Exportando...")
        #Cargamos el archivo
        wb = load_workbook('./recursos/template.xlsx')
        fecha = datetime.strftime(datetime.now(), '%Y-%m-%d')
        #indicamos la hoja
        ws = wb["presupuesto"]
        #Insertamos los datos en la cabecera
        ws['B2'] = f"Precios - Fecha: {fecha}"
        #Insertamos los productos
        inicio = 4
        for x in range(0, len(productos)):
            ws['B{a}'.format(a = str(inicio))] = productos[x][0]
            ws['D{a}'.format(a = str(inicio))] = productos[x][1]
            ws['E{a}'.format(a = str(inicio))] = productos[x][3]
            ws['G{a}'.format(a = str(inicio))] = productos[x][4]
            inicio += 1
        #Guarda el archivo
        wb.save('Presupuesto_{a}.xlsx'.format(a = fecha))
